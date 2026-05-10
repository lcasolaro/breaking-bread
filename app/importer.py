"""
Parses RICETTE IMPASTI.xlsx and seeds the DB.

Two sheet layouts exist:
 - "Napoletana/Teglia" style: Tipo Pizza in header row, toppings with kcal/100g column,
   variant data in (kcal, gr) column pairs starting at condimenti_col+2.
 - "Michele/Lioniello" style: variant names directly in idratazione row at col 8+,
   topping grams (no kcal) at those same columns.
"""

import os
import openpyxl
import app.db as db

EXCEL_PATH = os.getenv("EXCEL_PATH", os.path.join(os.path.dirname(os.path.dirname(__file__)), "RICETTE IMPASTI.xlsx"))

RECIPE_SHEETS = ["Pizza Napoletana", "Pizza In Teglia", "Ricetta Michele", "Ricetta Lioniello"]
TIMING_SHEET  = "Tempistiche"


# ── Helpers ──────────────────────────────────────────────────────────────────

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, str) and val.startswith('#'):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def find_row_with_text(rows, *texts, min_col=0, max_col=None):
    """Return index of first row that contains ALL given texts (anywhere in row)."""
    for i, row in enumerate(rows):
        cells = [str(c) for c in (row[min_col:max_col] if max_col else row[min_col:]) if c is not None]
        if all(any(t in c for c in cells) for t in texts):
            return i
    return -1


def find_col_in_row(row, text):
    """Return 0-based column index of first cell containing text."""
    for i, cell in enumerate(row):
        if cell is not None and isinstance(cell, str) and text.lower() in cell.lower():
            return i
    return -1


def scan_param(rows, label_text, col=0, value_col=1):
    """Find the value in value_col of the first row whose col contains label_text."""
    for row in rows:
        if row[col] is not None and isinstance(row[col], str) and label_text in row[col]:
            return row[value_col] if len(row) > value_col else None
    return None


# ── Recipe-sheet parsers ──────────────────────────────────────────────────────

def parse_napoletana_style(rows, sheet_name):
    """
    Handles Pizza Napoletana and Pizza In Teglia.
    'Tipo Pizza' header row → variant names every 2 cols.
    Condimenti block: (name, kcal/100g, kcal_v0, gr_v0, kcal_v1, gr_v1, ...)
    """
    params = {
        "hydration_pct": safe_float(scan_param(rows, "Idratazione")),
        "default_ball_g": safe_float(scan_param(rows, "Peso panetto")),
        "default_pieces": int(safe_float(scan_param(rows, "Numero pizze"), 1)),
        "salt_pct": safe_float(scan_param(rows, "Sale")),
        "biga_pct": safe_float(scan_param(rows, "BIGA")),
        "poolish_pct": safe_float(scan_param(rows, "POOLISH")),
        "autolisi_pct": safe_float(scan_param(rows, "AUTOLISI")),
        "base_flour_g": safe_float(scan_param(rows, "Farina totale")),
    }

    # Compute yeast % from CHIUSURA section if available
    chiusura_row = find_row_with_text(rows, "CHIUSURA IMPASTO")
    if chiusura_row >= 0 and params["base_flour_g"] > 0:
        for row in rows[chiusura_row:chiusura_row + 10]:
            if row[0] is not None and "Lievito" in str(row[0]) and row[1]:
                params["yeast_pct"] = round(safe_float(row[1]) / params["base_flour_g"] * 100, 3)
                break
    if "yeast_pct" not in params:
        params["yeast_pct"] = 0.0

    # Find 'Tipo Pizza' row → variant names
    tipo_pizza_row_idx = -1
    tipo_pizza_col = -1
    for i, row in enumerate(rows):
        col = find_col_in_row(row, "Tipo Pizza")
        if col >= 0:
            tipo_pizza_row_idx = i
            tipo_pizza_col = col
            break

    variant_names = []
    variant_gr_cols = []   # column index of 'gr' for each variant
    if tipo_pizza_row_idx >= 0:
        tipo_row = rows[tipo_pizza_row_idx]
        # Variant names appear every 2 cols starting at tipo_pizza_col+2
        j = tipo_pizza_col + 2
        while j < len(tipo_row):
            v = tipo_row[j]
            if v is not None and isinstance(v, str) and v.strip():
                variant_names.append(v.strip())
            j += 2

    # Find 'Condimenti' row
    condimenti_row_idx = -1
    condimenti_col = -1
    for i, row in enumerate(rows):
        col = find_col_in_row(row, "Condimenti")
        if col >= 0:
            condimenti_row_idx = i
            condimenti_col = col
            break

    toppings_by_variant = {n: [] for n in variant_names}
    if condimenti_row_idx >= 0 and condimenti_col >= 0 and variant_names:
        # gr cols: condimenti_col+3, condimenti_col+5, ...
        for vi in range(len(variant_names)):
            variant_gr_cols.append(condimenti_col + 3 + 2 * vi)
        kcal100_col = condimenti_col + 1

        STOP_LABELS = {"Numero Pizze", "Totale Ingredienti", "No. Fette", "No. Tranci", "Kcal/porzione", "Kcal/cena"}
        SKIP_LABELS = {"Farina"}

        # Toppings are the rows after the Condimenti header until 'Numero Pizze'
        for row in rows[condimenti_row_idx + 1:]:
            name_cell = row[condimenti_col] if condimenti_col < len(row) else None
            if name_cell is None or not isinstance(name_cell, str):
                continue
            name = name_cell.strip()
            if not name:
                continue
            if any(s in name for s in STOP_LABELS):
                break
            if name in SKIP_LABELS:
                continue
            kcal100 = safe_float(row[kcal100_col]) if kcal100_col < len(row) else None
            for vi, vname in enumerate(variant_names):
                gr_col = variant_gr_cols[vi]
                gr = safe_float(row[gr_col]) if gr_col < len(row) else 0
                if gr > 0:
                    toppings_by_variant[vname].append({
                        "name": name,
                        "quantity_g": gr,
                        "kcal_per100": kcal100 if kcal100 and kcal100 > 0 else None,
                    })

    return params, variant_names, toppings_by_variant


def parse_michele_style(rows, sheet_name):
    """
    Handles Ricetta Michele and Ricetta Lioniello.
    Variant names directly in idratazione row at col 8+.
    Toppings block: (name, ignored, gr_v0, gr_v1, ...) — no kcal/100g.
    """
    params = {
        "hydration_pct": safe_float(scan_param(rows, "Idratazione")),
        "default_ball_g": safe_float(scan_param(rows, "Peso panetto")),
        "default_pieces": int(safe_float(scan_param(rows, "Numero pizze"), 1)),
        "salt_pct": safe_float(scan_param(rows, "Sale")),
        "biga_pct": safe_float(scan_param(rows, "BIGA")),
        "poolish_pct": safe_float(scan_param(rows, "POOLISH")),
        "autolisi_pct": safe_float(scan_param(rows, "AUTOLISI")),
        "base_flour_g": safe_float(scan_param(rows, "Farina totale")),
        "yeast_pct": 0.0,
    }

    # Variant names: in the row containing "gr" at col 7, names at col 8+
    variant_names = []
    variant_start_col = 8
    for row in rows:
        if len(row) > 7 and row[7] == 'gr' and row[8] is not None:
            j = 8
            while j < len(row) and row[j] is not None and isinstance(row[j], str):
                variant_names.append(row[j].strip())
                j += 1
            variant_start_col = 8
            break

    # Find 'Condimenti' row
    condimenti_row_idx = -1
    condimenti_col = -1
    for i, row in enumerate(rows):
        col = find_col_in_row(row, "Condimenti")
        if col >= 0:
            condimenti_row_idx = i
            condimenti_col = col
            break

    toppings_by_variant = {n: [] for n in variant_names}
    if condimenti_row_idx >= 0 and condimenti_col >= 0 and variant_names:
        # For each topping row: name at condimenti_col, grams at variant_start_col+i
        for row in rows[condimenti_row_idx + 1:]:
            name_cell = row[condimenti_col] if condimenti_col < len(row) else None
            if name_cell is None or not isinstance(name_cell, str):
                continue
            name = name_cell.strip()
            if not name:
                continue

            for vi, vname in enumerate(variant_names):
                gr_col = variant_start_col + vi
                gr = safe_float(row[gr_col]) if gr_col < len(row) else 0
                if gr > 0:
                    toppings_by_variant[vname].append({
                        "name": name,
                        "quantity_g": gr,
                        "kcal_per100": None,
                    })

    return params, variant_names, toppings_by_variant


def parse_timings(ws):
    """Extract timing guide sections from the Tempistiche sheet."""
    guides = []
    current_name = None
    current_lines = []

    for row in ws.iter_rows(values_only=True):
        non_empty = [c for c in row if c is not None]
        if not non_empty:
            continue
        first = str(non_empty[0]).strip()
        if not first:
            continue

        # Section header: bold/standalone title text (all-caps or short header)
        # Heuristic: if first cell has no other meaningful cols and is a title-like string
        other_cols = [c for c in row[1:] if c is not None]
        if not other_cols and len(first) < 60:
            # Likely a section header
            if current_name and current_lines:
                guides.append({"name": current_name, "content": "\n".join(current_lines)})
            current_name = first
            current_lines = []
        else:
            if current_name:
                line_parts = [str(c).strip() for c in row if c is not None]
                current_lines.append("  ".join(line_parts))

    if current_name and current_lines:
        guides.append({"name": current_name, "content": "\n".join(current_lines)})

    return guides


# ── Main import function ──────────────────────────────────────────────────────

def import_excel(source=None, reset: bool = False, only_names=None) -> dict:
    """
    Import recipes from Excel into the DB.
    source can be a file path (str) or a BytesIO object.
    If reset=True, clears all existing recipes first.
    Returns a summary dict.
    """
    if source is None:
        source = EXCEL_PATH

    errors = []
    recipes_added = 0
    variants_added = 0
    toppings_added = 0
    timing_added = 0

    if isinstance(source, (str, os.PathLike)):
        if not os.path.exists(source):
            return {"ok": False, "error": f"File non trovato: {source}"}
        path_label = os.path.basename(source)
    else:
        path_label = getattr(source, 'name', 'upload.xlsx')

    try:
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Errore apertura file: {e}"}

    if 'Ricette' in wb.sheetnames:
        return _import_from_template(wb, reset, only_names=only_names)

    if reset:
        # Drop and re-create all data (preserve schema)
        existing = db.get_recipes()
        for r in existing:
            db.delete_recipe(r["id"])

    # ── Parse recipe sheets ──────────────────────────────────────────────────
    napoletana_style = {"Pizza Napoletana", "Pizza In Teglia"}

    for sheet_name in RECIPE_SHEETS:
        if sheet_name not in wb.sheetnames:
            errors.append(f"Sheet '{sheet_name}' non trovata")
            continue

        if only_names is not None and sheet_name not in only_names:
            continue

        # Skip if recipe already exists
        existing_names = {r["name"] for r in db.get_recipes()}
        if sheet_name in existing_names:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        try:
            if sheet_name in napoletana_style:
                params, variant_names, toppings_by_variant = parse_napoletana_style(rows, sheet_name)
            else:
                params, variant_names, toppings_by_variant = parse_michele_style(rows, sheet_name)
        except Exception as e:
            errors.append(f"Errore parsing '{sheet_name}': {e}")
            continue

        if not params.get("base_flour_g") or not params.get("default_pieces"):
            errors.append(f"Parametri mancanti in '{sheet_name}', saltata")
            continue

        recipe_id = db.create_recipe({
            "name": sheet_name,
            "description": None,
            "base_flour_g": params.get("base_flour_g", 1000),
            "default_pieces": params.get("default_pieces", 1),
            "default_ball_g": params.get("default_ball_g", 250),
            "hydration_pct": params.get("hydration_pct", 70),
            "salt_pct": params.get("salt_pct", 2.5),
            "yeast_pct": params.get("yeast_pct", 0),
            "biga_pct": params.get("biga_pct", 0),
            "poolish_pct": params.get("poolish_pct", 0),
            "autolisi_pct": params.get("autolisi_pct", 0),
            "extra_ingredients": [],
            "notes": None,
            "sort_order": RECIPE_SHEETS.index(sheet_name),
        })
        recipes_added += 1

        for vi, vname in enumerate(variant_names):
            variant_id = db.create_variant(recipe_id, vname, sort_order=vi)
            variants_added += 1
            for topping in toppings_by_variant.get(vname, []):
                db.create_topping(variant_id, topping)
                toppings_added += 1

    # ── Parse timing sheet ───────────────────────────────────────────────────
    if TIMING_SHEET in wb.sheetnames:
        existing_guides = {g["name"] for g in db.get_timing_guides()}
        ws = wb[TIMING_SHEET]
        guides = parse_timings(ws)
        for i, guide in enumerate(guides):
            if guide["name"] not in existing_guides:
                db.create_timing_guide(guide["name"], guide["content"], sort_order=i)
                timing_added += 1

    db.log_import(
        path_label,
        recipes_added,
        f"varianti: {variants_added}, topping: {toppings_added}, guide: {timing_added}"
    )

    return {
        "ok": True,
        "recipes_added": recipes_added,
        "variants_added": variants_added,
        "toppings_added": toppings_added,
        "timing_guides_added": timing_added,
        "errors": errors,
    }


def _import_from_template(wb, reset: bool, only_names=None):
    ws = wb['Ricette']
    ws_var = wb['Varianti'] if 'Varianti' in wb.sheetnames else None

    recipes_added = 0; variants_added = 0; toppings_added = 0; errors = []
    only_set = set(only_names) if only_names is not None else None

    if reset:
        existing_recipes = db.get_recipes()
        for r in existing_recipes:
            if only_set is None or r['name'] in only_set:
                db.delete_recipe(r["id"])

    # Read header row to get column mapping
    headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}

    def col(row_vals, name, default=None):
        idx = headers.get(name)
        if idx is None: return default
        v = row_vals[idx] if idx < len(row_vals) else None
        return v if v is not None else default

    def sf(v, d=0.0):
        try: return float(v) if v is not None else d
        except: return d

    def si(v, d=0):
        try: return int(float(v)) if v is not None else d
        except: return d

    # Read existing recipe names to avoid duplicates (refresh after potential reset)
    existing = {r['name'] for r in db.get_recipes()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        row = list(row)
        name = str(col(row, 'Nome', '')).strip()
        if not name: continue
        if only_set is not None and name not in only_set: continue
        if name in existing:
            errors.append(f'"{name}" già presente, saltata')
            continue

        pieces = si(col(row, 'Num. Panetti', 6))
        ball_g = sf(col(row, 'Peso Panetto (g)', 255))
        hydrat = sf(col(row, 'Idratazione (%)', 65))
        flour  = pieces * ball_g / (1 + hydrat / 100)

        data = {
            'name': name,
            'description': str(col(row, 'Descrizione', '') or '').strip() or None,
            'base_flour_g': round(flour, 1),
            'default_pieces': pieces,
            'default_ball_g': ball_g,
            'hydration_pct': hydrat,
            'biga_pct':    sf(col(row, 'BIGA (%)', 0)),
            'biga_hydration_pct': sf(col(row, 'Idrat. BIGA (%)', 44)),
            'biga_yeast_pct':     sf(col(row, 'Lievito BIGA (% far. BIGA)', 0.5)),
            'poolish_pct': sf(col(row, 'POOLISH (%)', 0)),
            'poolish_yeast_pct':  sf(col(row, 'Lievito POOLISH (% far. POOLISH)', 0.1)),
            'autolisi_pct':sf(col(row, 'AUTOLISI (%)', 0)),
            'autolisi_water_pct': sf(col(row, 'Acqua AUTOLISI (%, 0=idrat.)', 0)),
            'salt_pct':    sf(col(row, 'Sale (% far. tot.)', 2.5)),
            'yeast_pct':   sf(col(row, 'Lievito tot. (% far. tot.)', 1.0)),
            'malto_pct': sf(col(row, 'Malto Diastasico (% biga+poolish)', 0)),
            'carbone_pct': 0,
            'olio_pct':    sf(col(row, 'Olio (% far. tot.)', 0)),
            'extra_ingredients': [],
            'notes': str(col(row, 'Note', '') or '').strip() or None,
            'sort_order': recipes_added * 10,
        }
        rid = db.create_recipe(data)
        existing.add(name)
        recipes_added += 1

        # Read variants for this recipe from Varianti sheet
        if ws_var:
            var_headers = {cell.value: cell.column - 1 for cell in ws_var[1] if cell.value}
            current_variants = {}
            for vrow in ws_var.iter_rows(min_row=2, values_only=True):
                if not any(vrow): continue
                vrow = list(vrow)
                def vc(n, d=None):
                    i = var_headers.get(n)
                    return vrow[i] if i is not None and i < len(vrow) and vrow[i] is not None else d
                rec_name = str(vc('Ricetta', '') or '').strip()
                if rec_name != name: continue
                var_name = str(vc('Variante', '') or '').strip()
                if not var_name: continue
                ing_name = str(vc('Ingrediente', '') or '').strip()
                if not ing_name: continue
                if var_name not in current_variants:
                    vid = db.create_variant(rid, var_name, len(current_variants) * 10)
                    current_variants[var_name] = vid
                    variants_added += 1
                vid = current_variants[var_name]
                db.create_topping(vid, {
                    'name': ing_name,
                    'quantity_g': sf(vc('g/pizza', 0)),
                    'kcal_per100': sf(vc('kcal/100g')) if vc('kcal/100g') else None,
                    'protein_per100': sf(vc('Proteine/100g')) if vc('Proteine/100g') else None,
                    'carbs_per100': sf(vc('Carboidrati/100g')) if vc('Carboidrati/100g') else None,
                    'fat_per100': sf(vc('Grassi/100g')) if vc('Grassi/100g') else None,
                    'sort_order': toppings_added,
                })
                toppings_added += 1

    return {
        'ok': True, 'format': 'template',
        'recipes_added': recipes_added, 'variants_added': variants_added,
        'toppings_added': toppings_added, 'timing_guides_added': 0, 'errors': errors
    }


def preview_excel_import(source) -> dict:
    """Parse Excel and return list of recipe names found, without importing."""
    try:
        if hasattr(source, 'seek'):
            source.seek(0)
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Errore apertura file: {e}"}

    existing = {r['name'] for r in db.get_recipes()}

    if 'Ricette' in wb.sheetnames:
        ws = wb['Ricette']
        headers = {cell.value: cell.column - 1 for cell in ws[1] if cell.value}
        recipes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row = list(row)
            idx = headers.get('Nome')
            name = str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ''
            if name:
                recipes.append({'name': name, 'already_exists': name in existing})
        return {'ok': True, 'format': 'template', 'recipes': recipes}
    else:
        recipes = [
            {'name': sn, 'already_exists': sn in existing}
            for sn in RECIPE_SHEETS if sn in wb.sheetnames
        ]
        return {'ok': True, 'format': 'legacy', 'recipes': recipes}


def export_to_excel(recipe_ids=None) -> "openpyxl.Workbook":
    """Export all recipes and variants from DB into the importable template format."""
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # ── Sheet 1: Istruzioni ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Istruzioni"
    ws.column_dimensions['A'].width = 72
    ws['A1'] = '🍕 BREAKING BREAD — Esportazione Ricette'
    ws['A1'].font = Font(bold=True, size=14, color='C8550A')
    ws['A2'] = 'Questo file può essere reimportato tramite il pulsante "Importa da Excel" nell\'app.'
    ws['A3'] = 'Le colonne NON devono essere rinominate. I numeri decimali usano il punto (es. 0.5).'

    # ── Sheet 2: Ricette ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ricette")
    recipe_headers = [
        'Nome', 'Descrizione',
        'Num. Panetti', 'Peso Panetto (g)', 'Idratazione (%)',
        'BIGA (%)', 'Idrat. BIGA (%)', 'Lievito BIGA (% far. BIGA)',
        'POOLISH (%)', 'Lievito POOLISH (% far. POOLISH)',
        'AUTOLISI (%)', 'Acqua AUTOLISI (%, 0=idrat.)',
        'Sale (% far. tot.)', 'Lievito tot. (% far. tot.)',
        'Malto Diastasico (% biga+poolish)', 'Olio (% far. tot.)', 'Note',
    ]
    col_widths = [28, 35, 14, 16, 14, 9, 14, 24, 12, 26, 12, 22, 18, 20, 26, 16, 35]
    for c, (h, w) in enumerate(zip(recipe_headers, col_widths), 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='C8550A')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws2.column_dimensions[ws2.cell(row=1, column=c).column_letter].width = w
    ws2.row_dimensions[1].height = 42

    recipes = db.get_recipes()
    if recipe_ids is not None:
        ids_set = set(recipe_ids)
        recipes = [r for r in recipes if r['id'] in ids_set]
    for i, r in enumerate(recipes, 2):
        vals = [
            r['name'],
            r.get('description') or '',
            r['default_pieces'],
            r['default_ball_g'],
            r['hydration_pct'],
            r['biga_pct'],
            r.get('biga_hydration_pct', 44),
            r.get('biga_yeast_pct', 0.5),
            r['poolish_pct'],
            r.get('poolish_yeast_pct', 0.1),
            r['autolisi_pct'],
            r.get('autolisi_water_pct', 0),
            r['salt_pct'],
            r['yeast_pct'],
            r.get('malto_pct', 0),
            r.get('olio_pct', 0),
            r.get('notes') or '',
        ]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=i, column=c, value=v)

    # ── Sheet 3: Varianti ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Varianti")
    variant_headers = ['Ricetta', 'Variante', 'Ingrediente', 'g/pizza', 'kcal/100g',
                       'Proteine/100g', 'Carboidrati/100g', 'Grassi/100g']
    widths3 = [28, 20, 28, 10, 12, 14, 16, 12]
    for c, (h, w) in enumerate(zip(variant_headers, widths3), 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a7fa8')
        cell.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[ws3.cell(row=1, column=c).column_letter].width = w

    row_idx = 2
    for r in recipes:  # already filtered by recipe_ids above
        full = db.get_recipe(r['id'])
        for variant in full.get('variants', []):
            for topping in variant.get('toppings', []):
                ws3.cell(row=row_idx, column=1, value=r['name'])
                ws3.cell(row=row_idx, column=2, value=variant['name'])
                ws3.cell(row=row_idx, column=3, value=topping['name'])
                ws3.cell(row=row_idx, column=4, value=topping['quantity_g'])
                ws3.cell(row=row_idx, column=5, value=topping.get('kcal_per100'))
                ws3.cell(row=row_idx, column=6, value=topping.get('protein_per100'))
                ws3.cell(row=row_idx, column=7, value=topping.get('carbs_per100'))
                ws3.cell(row=row_idx, column=8, value=topping.get('fat_per100'))
                row_idx += 1

    return wb


def create_import_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # Sheet 1: Istruzioni
    ws = wb.active
    ws.title = "Istruzioni"
    ws.column_dimensions['A'].width = 72
    ws['A1'] = '🍕 BREAKING BREAD — Template Import Ricette'
    ws['A1'].font = Font(bold=True, size=14, color='C8550A')
    instructions = [
        '',
        '📋 ISTRUZIONI:',
        '',
        '1. Compila il foglio "Ricette" — ogni riga = una ricetta.',
        '   Le percentuali si inseriscono come numeri interi o decimali (es. 65 per 65%).',
        '',
        '2. Compila il foglio "Varianti" — ogni riga = un condimento di una variante.',
        '   La colonna "Ricetta" deve corrispondere ESATTAMENTE al nome nel foglio Ricette.',
        '   La colonna "Variante" raggruppa i condimenti (es. Margherita, Marinara...).',
        '',
        '3. Salva il file con il nome "ricette.xlsx" nella cartella dell\'app.',
        '   Poi clicca "Importa da Excel" nell\'app.',
        '',
        '⚠️ NOTE IMPORTANTI:',
        '• Le colonne con intestazione non devono essere rinominate.',
        '• Le ricette con lo stesso nome vengono saltate (non duplicate).',
        '• I numeri decimali usano il punto come separatore (es. 0.5, non 0,5).',
        '• Lascia 0 nei campi che non usi (es. BIGA=0 se non usi la biga).',
        '• Il carbone vegetale è fisso a 7g/kg di farina — non serve indicarlo.',
        '',
        '📐 CALCOLO IMPASTO:',
        '• Impasto totale = Num.Panetti × Peso Panetto',
        '• Farina = Impasto totale ÷ (1 + Idratazione/100)',
        '• Acqua totale = Farina × Idratazione/100',
        '• Le altre percentuali (BIGA, POOLISH, ecc.) sono sulla farina totale.',
    ]
    for i, line in enumerate(instructions):
        ws.cell(row=i+2, column=1, value=line)

    # Sheet 2: Ricette
    ws2 = wb.create_sheet("Ricette")
    headers = [
        'Nome', 'Descrizione',
        'Num. Panetti', 'Peso Panetto (g)', 'Idratazione (%)',
        'BIGA (%)', 'Idrat. BIGA (%)', 'Lievito BIGA (% far. BIGA)',
        'POOLISH (%)', 'Lievito POOLISH (% far. POOLISH)',
        'AUTOLISI (%)', 'Acqua AUTOLISI (%, 0=idrat.)',
        'Sale (% far. tot.)', 'Lievito tot. (% far. tot.)',
        'Malto Diastasico (% biga+poolish)', 'Olio (% far. tot.)', 'Note'
    ]
    col_widths = [28, 35, 14, 16, 14, 9, 14, 24, 12, 26, 12, 22, 18, 20, 26, 16, 35]
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='C8550A')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        col_letter = ws2.cell(row=1, column=c).column_letter
        ws2.column_dimensions[col_letter].width = w
    ws2.row_dimensions[1].height = 42
    # Example row
    example = ['🍕 Pizza Napoletana', 'Impasto napoletano contemporaneo',
               6, 250, 65, 20, 44, 0.5, 10, 0.1, 0, 0, 2.5, 1.0, 0, 0, 'Lunga maturazione in frigo']
    for c, v in enumerate(example, 1):
        ws2.cell(row=2, column=c, value=v)

    # Sheet 3: Varianti
    ws3 = wb.create_sheet("Varianti")
    headers3 = ['Ricetta', 'Variante', 'Ingrediente', 'g/pizza', 'kcal/100g',
                'Proteine/100g', 'Carboidrati/100g', 'Grassi/100g']
    widths3 = [28, 20, 28, 10, 12, 14, 16, 12]
    for c, (h, w) in enumerate(zip(headers3, widths3), 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a7fa8')
        cell.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[ws3.cell(row=1, column=c).column_letter].width = w
    ex_var = [
        ('🍕 Pizza Napoletana','Margherita','Pomodoro San Marzano', 80, 18, 1.0, 3.5, 0.2),
        ('🍕 Pizza Napoletana','Margherita','Mozzarella fior di latte', 80, 242, 17.1, 2.7, 18.3),
        ('🍕 Pizza Napoletana','Margherita','Basilico fresco', 5, 22, 3.2, 1.1, 0.6),
        ('🍕 Pizza Napoletana','Marinara','Pomodoro San Marzano', 100, 18, 1.0, 3.5, 0.2),
        ('🍕 Pizza Napoletana','Marinara','Aglio', 5, 149, 6.4, 33.1, 0.5),
    ]
    for r, row in enumerate(ex_var, 2):
        for c, v in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=v)

    return wb
